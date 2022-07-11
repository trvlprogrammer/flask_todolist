
from flask import flash, render_template,redirect, url_for, request, send_file, current_app
from app.todolist import bp
from app.todolist.forms import PostForm, TagForm, UploadForm
from app.models import Post, Tag
from app import create_app, db
from flask_login import current_user, login_required
import openpyxl
import os
from app.utils import mail
from werkzeug.utils import secure_filename


@bp.route('/',methods=['GET', 'POST'])
@bp.route('/index', methods=['GET', 'POST'])
@login_required
def index():    
    form = PostForm()
    form.set_choices(current_user)    
    if form.validate_on_submit():
        body = form.body.data
        date_todo = form.date_todo.data
        if date_todo:
            date_todo = date_todo.utcnow()
        tags  = db.session.query(Tag).filter(Tag.id.in_(form.tags.data)).all()
        post = Post(body=body, tags=tags, date_todo= date_todo,author=current_user)
        db.session.add(post)
        db.session.commit()

        flash('Todo has been added', 'success')
        return redirect(url_for('todolist.index'))
    page = request.args.get('page', 1, type=int)
    posts = Post.query.filter_by(user_id=current_user.id,active=True).paginate(page, current_app.config['POSTS_PER_PAGE'])
    next_url = url_for('todolist.index', page=posts.next_num) if posts.has_next else None        
    prev_url = url_for('todolist.index', page=posts.prev_num) if posts.has_prev else None                
    return render_template('index.html',title='Todo', form=form, posts=posts.items, \
     next_url=next_url,prev_url=prev_url,uploadform=UploadForm(),name="Todolist")

@bp.route('/tags', methods=['GET', 'POST'])
@login_required
def tag_view():
    form = TagForm()    
    if form.validate_on_submit():
        tag = Tag(name=form.name.data,tag_user=current_user)
        db.session.add(tag)
        db.session.commit()
        flash('Tag has been added', 'success')
        return redirect(url_for('todolist.tag_view'))
    
    tags = current_user.tags.all()
    return render_template('tags.html',title='Tags', form=form, tags=tags)

@bp.route('/todo/archive', methods=['GET', 'POST'])
@login_required
def todo_archive():
    page = request.args.get('page', 1, type=int)
    posts = Post.query.filter_by(user_id=current_user.id,active=False).paginate(page, current_app.config['POSTS_PER_PAGE'])
    next_url = url_for('todolist.todo_archive', page=posts.next_num) if posts.has_next else None        
    prev_url = url_for('todolist.todo_archive', page=posts.prev_num) if posts.has_prev else None  
    return render_template('index.html',title='Archive', posts=posts.items, next_url=next_url,prev_url=prev_url, \
    uploadform=UploadForm(),name="Archive")

@bp.route('/todo/set_active/<todo_id>/<active>', methods=['POST'])
@login_required
def todo_active_inactive(todo_id,active):
    post = Post.query.get(todo_id)
    post.active = eval(active)
    db.session.commit()
    return redirect(request.referrer)

@bp.route('/todo/export', methods=['GET'])
@login_required
def export_todo():     
    file_path = export_todo_file(current_user)
    return send_file(file_path)

@bp.route('/todo/<todo_id>', methods=['POST'])
@login_required
def delete_todo(todo_id):
    Post.query.filter_by(id=todo_id).delete()
    db.session.commit()
    return redirect(request.referrer)

@bp.route('/todo/export_email', methods=['GET'])
@login_required
def export_todo_email():    
    file_path = export_todo_file(current_user)           
    with current_app.open_resource(file_path) as fp:
        attachment = {
            'file_name' : 'todo_export.xlsx',
            'content_type' : 'file/xlsx',
            'data' : fp.read()
        }
    html_body = render_template('email/export_todo.html',user=current_user)
    mail.send_email('Export Todo', current_app.config['ADMINS'][0], [current_user.email], None, html_body,attachment)
    return redirect(request.referrer)

def export_todo_file(current_user):
    file_name = 'todo_export.xlsx'
    file_path = os.path.join(current_app.config['EXPORT_PATH'], file_name)
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    sheet["A1"]= 'Descripton'
    sheet["B1"]= 'Active'
    i = 1
    for todo in current_user.posts.all():
        i+=1
        sheet['A'+str(i)]= todo.body
        sheet['B'+str(i)]= todo.active
    wb.save(file_path)
    return file_path

@bp.route('/todo/import', methods=['POST'])
@login_required
def import_todo():
    if 'file' not in request.files:
        flash('No file part','danger')
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        flash('No selected file','danger')
        return redirect(request.referrer)
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(current_app.config['UPLOAD_PATH'], filename)
        file.save(file_path)
        wb_obj = openpyxl.load_workbook(file_path)
        ws = wb_obj.active
        data_cell = ws["A:B"]            
        i = 1
        while i < len(data_cell[0]):
            post = Post(body=data_cell[0][i].value,active=data_cell[1][i].value,author=current_user)
            db.session.add(post)
            db.session.commit()
            i += 1
        return redirect(request.referrer)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in current_app.config['ALLOWED_EXTENSIONS'] 
           
@bp.route('/tag/<tag_id>', methods=['POST'])
@login_required
def delete_tag(tag_id):
    Tag.query.filter_by(id=tag_id).delete()
    db.session.commit()
    return redirect(request.referrer)


@bp.route('/todo/download-template-import', methods=['GET'])
@login_required
def download_template_import():     
    
    return send_file(os.path.join('templates/import','todolist_template_import.xlsx'))
